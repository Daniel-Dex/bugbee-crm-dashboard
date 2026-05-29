from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .transform import money, parent_sku, pick, split_product_variant


def latest_file(folder: Path, pattern: str, fallback: Path) -> Path:
    if fallback.exists():
        return fallback
    if folder.exists():
        matches = sorted(folder.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
        if matches:
            return matches[0]
    return fallback


BASE_MAGAZORD = latest_file(
    Path("C:/Users/Daniel/OneDrive/Documentos/Clientologia/Bases de Análise Magazord/2026"),
    "Base Análise Magazord - Vendas de Maio*.xlsx",
    Path("C:/Users/Daniel/OneDrive/Documentos/Clientologia/Bases de Análise Magazord/2026/Base Análise Magazord - Vendas de Maio de 01-05 até 15-05.xlsx"),
)

PECAS = latest_file(
    Path("C:/Users/Daniel/OneDrive/Documentos/Clientologia/Estoque/2026"),
    "Tabela Peças - Vendas Maio*.xlsx",
    Path("C:/Users/Daniel/OneDrive/Documentos/Clientologia/Estoque/2026/Tabela Peças - Vendas Maio de 01-05 até 15-05.xlsx"),
)


def clean(value):
    if isinstance(value, str) and value.startswith("#"):
        return None
    return value


def _rows_by_header(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {str(h): clean(v) for h, v in zip(headers, values) if h not in (None, "")}
        if row:
            rows.append(row)
    return rows


def _price_rank(row: dict[str, Any]) -> tuple[int, int]:
    table_name = str(pick(row, "Tabela de Preços Nome", default="") or "").strip().lower()
    marketplace_id = money(pick(row, "Marketplace Id", default=0))
    return (0 if table_name == "padrão" else 1, 0 if marketplace_id == 0 else 1)


def _apply_price_row(lookup: dict[str, dict[str, Any]], row: dict[str, Any]) -> None:
    sku = pick(row, "Produto Código", "Produto/Derivação Código Der.")
    if not sku:
        return

    old_price = money(pick(row, "Preço Antigo", "Preço cheio", "Preço Cheio"))
    sale_price = money(pick(row, "Preço Venda", "Preço venda", "Preço Venda", "Valor Unitário"))
    full_price = old_price if old_price > 0 else sale_price
    if full_price <= 0:
        return

    key = str(sku)
    rank = _price_rank(row)
    info = lookup.setdefault(key, {"sku": key, "sku_pai": parent_sku(key)})
    current_rank = info.get("_preco_rank", (99, 99))
    if info.get("preco_cheio") and current_rank <= rank:
        return

    info["preco_cheio"] = full_price
    info["preco_tabela_venda"] = sale_price
    info["preco_antigo"] = old_price
    info["_preco_rank"] = rank


def load_reference_lookup() -> dict[str, dict[str, Any]]:
    lookup: dict[str, dict[str, Any]] = {}

    base_rows = _rows_by_header(BASE_MAGAZORD, "Estoque Magazord") + _rows_by_header(BASE_MAGAZORD, "Descontos")
    for row in base_rows:
        sku = pick(row, "Produto/Derivação Código Der.", "Produto Código")
        if not sku:
            continue
        desc = pick(row, "Produto/Derivação Derivação")
        produto, cor_guess, tamanho_guess = split_product_variant(desc)
        info = {
            "sku": str(sku),
            "sku_pai": parent_sku(sku),
            "produto": produto or desc,
            "descricao_completa": desc,
            "cor": pick(row, "Cor", default=cor_guess),
            "tamanho": pick(row, "Grade", default=tamanho_guess),
            "categoria": pick(row, "Categoria", "Categoria Principal"),
            "genero": pick(row, "Gênero", "Genero", "Nome Grupo Produto"),
            "colecao": pick(row, "Coleção"),
            "linha_produto": pick(row, "Linha de Produto", "Nome Linha"),
            "custo_medio": money(pick(row, "Custo Médio de Estoque", "Custo Médio", "Custo unitário", "Custo")),
            "quantidade_comprada": money(pick(row, "Quantidade Prevista Entrada", "Quantidade Comprada")),
        }
        lookup[str(sku)] = {k: v for k, v in info.items() if v not in (None, "")}

    price_rows = (
        _rows_by_header(BASE_MAGAZORD, "Planilha7")
        + _rows_by_header(BASE_MAGAZORD, "Tabela Preço Cheio")
        + _rows_by_header(PECAS, "Tabela de Preço 2")
        + _rows_by_header(PECAS, "Planilha1")
    )
    for row in price_rows:
        _apply_price_row(lookup, row)

    peca_rows = _rows_by_header(PECAS, "Maio") + _rows_by_header(PECAS, "entradas e devoluções")
    for row in peca_rows:
        sku = pick(row, "Produto Código")
        if not sku:
            continue
        info = lookup.setdefault(str(sku), {"sku": str(sku), "sku_pai": parent_sku(sku)})
        if not info.get("colecao"):
            info["colecao"] = pick(row, "Caracteristica")
        if not info.get("categoria"):
            info["categoria"] = pick(row, "Categoria Principal")
        if not info.get("custo_medio"):
            info["custo_medio"] = money(pick(row, "Custo Médio"))
        if not info.get("preco_cheio"):
            preco_venda = money(pick(row, "Valor Unitário"))
            if preco_venda:
                info["preco_cheio"] = preco_venda

    for info in lookup.values():
        info.pop("_preco_rank", None)
    return lookup
