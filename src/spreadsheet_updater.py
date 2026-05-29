from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="E8546A")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ACCOUNTING_FMT = '"R$" #,##0.00'
INTEGER_FMT = '#,##0'
NUMBER_2_FMT = '#,##0.00'
PERCENT_FMT = '0.00%'
FORCE_MONEY_HEADERS = ("valor_desconto", "valor desconto")

MONEY_KEYS = (
    "faturamento",
    "receita",
    "venda",
    "valor",
    "preco",
    "preço",
    "custo",
    "devolu",
    "ticket",
    "margem",
    "cpv",
    "frete",
)
PERCENT_KEYS = ("pct", "%", "percent", "desconto", "taxa", "roas", "cvr")
INTEGER_KEYS = ("quantidade", "qtd", "pedidos", "clientes", "itens", "peças", "pecas", "estoque")
DECIMAL_KEYS = ("ticket médio", "peças por pedido", "pecas por pedido", "pa")


def normalize_label(value) -> str:
    return str(value or "").strip().lower()


def number_format_for_header(header: str) -> str | None:
    text = normalize_label(header)
    if text in FORCE_MONEY_HEADERS:
        return ACCOUNTING_FMT
    if any(key in text for key in PERCENT_KEYS):
        return PERCENT_FMT
    if any(key in text for key in MONEY_KEYS):
        return ACCOUNTING_FMT
    if any(key in text for key in DECIMAL_KEYS):
        return NUMBER_2_FMT
    if any(key in text for key in INTEGER_KEYS):
        return INTEGER_FMT
    return None


def number_format_for_indicator(indicator: str) -> str | None:
    text = normalize_label(indicator)
    if "desconto" in text or "%" in text or "taxa" in text:
        return PERCENT_FMT
    if any(key in text for key in ("faturamento", "receita", "venda", "valor", "ticket", "devolu")):
        return ACCOUNTING_FMT
    if any(key in text for key in ("peças por pedido", "pecas por pedido")):
        return NUMBER_2_FMT
    if any(key in text for key in ("pedidos", "clientes", "itens")):
        return INTEGER_FMT
    return None


def autofit_columns(ws) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = 0
        for cell in column_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 46)


def format_worksheet(ws) -> None:
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    value_col = None
    indicator_col = None
    for idx, header in enumerate(headers, 1):
        text = normalize_label(header)
        if text == "valor":
            value_col = idx
        elif text == "indicador":
            indicator_col = idx

    for col, header in enumerate(headers, 1):
        header_format = number_format_for_header(str(header or ""))
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, col)
            if not isinstance(cell.value, (int, float)):
                continue
            number_format = header_format
            if value_col == col and indicator_col:
                number_format = number_format_for_indicator(ws.cell(row, indicator_col).value) or header_format
            if number_format:
                cell.number_format = number_format
            cell.alignment = Alignment(horizontal="right")

    if ws.title == "Comparativo Semanal":
        indicator_col = next((idx for idx, header in enumerate(headers, 1) if normalize_label(header) == "indicador"), None)
        for row in range(2, ws.max_row + 1):
            indicator = ws.cell(row, indicator_col).value if indicator_col else ""
            value_format = number_format_for_indicator(indicator) or NUMBER_2_FMT
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row, col)
                if not isinstance(cell.value, (int, float)):
                    continue
                header_text = normalize_label(header)
                if header_text == "pct":
                    cell.number_format = PERCENT_FMT
                elif header_text == "abs":
                    cell.number_format = value_format
                elif header_text in ("atual", "semana_anterior", "ano_anterior"):
                    cell.number_format = value_format

    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center")
    autofit_columns(ws)


def write_table(ws, rows: list[dict], start_row: int = 1) -> int:
    if not rows:
        ws.cell(start_row, 1, "sem_dados")
        return start_row + 1
    headers = list(rows[0].keys())
    for col, header in enumerate(headers, 1):
        cell = ws.cell(start_row, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for r, row in enumerate(rows, start_row + 1):
        for c, header in enumerate(headers, 1):
            ws.cell(r, c, row.get(header))
    format_worksheet(ws)
    return start_row + len(rows) + 1


def latest_file(folder: Path, pattern: str, fallback: Path) -> Path:
    if fallback.exists():
        return fallback
    if folder.exists():
        matches = sorted(folder.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
        if matches:
            return matches[0]
    return fallback


SOURCE_WORKBOOKS = {
    "Modelo Base Magazord": latest_file(
        Path("C:/Users/Daniel/OneDrive/Documentos/Clientologia/Bases de Análise Magazord/2026"),
        "Base Análise Magazord - Vendas de Maio*.xlsx",
        Path("C:/Users/Daniel/OneDrive/Documentos/Clientologia/Bases de Análise Magazord/2026/Base Análise Magazord - Vendas de Maio de 01-05 até 15-05.xlsx"),
    ),
    "Modelo Clientes": Path("C:/Users/Daniel/OneDrive/Documentos/Clientologia/Clientes/Controle de clientes - 2026.xlsx"),
    "Modelo Peças": latest_file(
        Path("C:/Users/Daniel/OneDrive/Documentos/Clientologia/Estoque/2026"),
        "Tabela Peças - Vendas Maio*.xlsx",
        Path("C:/Users/Daniel/OneDrive/Documentos/Clientologia/Estoque/2026/Tabela Peças - Vendas Maio de 01-05 até 15-05.xlsx"),
    ),
}


def copy_preview_sheet(wb: Workbook, source_path: Path, sheet_name: str, target_name: str, max_rows: int = 80, max_cols: int = 25) -> None:
    if not source_path.exists():
        return
    source = load_workbook(source_path, read_only=True, data_only=False)
    if sheet_name not in source.sheetnames:
        return
    ws_src = source[sheet_name]
    ws = wb.create_sheet(target_name[:31])
    for r, row in enumerate(ws_src.iter_rows(min_row=1, max_row=min(ws_src.max_row, max_rows), max_col=min(ws_src.max_column, max_cols)), 1):
        for c, src_cell in enumerate(row, 1):
            ws.cell(r, c, src_cell.value)
    ws.freeze_panes = "A2"


def build_weekly_workbook(output_path: Path, data: dict) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    tabs = {
        "Base API Tratada": data.get("orders", []),
        "Base Consolidada": data.get("items", []),
        "Devolucoes": data.get("return_items", []),
        "Indicadores Semanais": data.get("weekly_indicator_rows", []),
        "Comparativo Semanal": data.get("comparison_rows", []),
        "Produtos e Categorias": data.get("product_category_rows", []),
        "Estoque e Giro": data.get("stock_rows", []),
        "Site": data.get("site_rows", []),
        "Marketplace": data.get("marketplace_rows", []),
        "Dados para Apresentacao": data.get("presentation_rows", []),
    }
    for name, rows in tabs.items():
        ws = wb.create_sheet(name[:31])
        write_table(ws, rows)
        ws.freeze_panes = "A2"
    copy_preview_sheet(wb, SOURCE_WORKBOOKS["Modelo Base Magazord"], "base analise", "Modelo base analise")
    copy_preview_sheet(wb, SOURCE_WORKBOOKS["Modelo Base Magazord"], "DRE", "Modelo DRE")
    copy_preview_sheet(wb, SOURCE_WORKBOOKS["Modelo Clientes"], "CAC", "Modelo CAC")
    copy_preview_sheet(wb, SOURCE_WORKBOOKS["Modelo Peças"], "Relatório", "Modelo Peças")
    for ws in wb.worksheets:
        format_worksheet(ws)
    wb.save(output_path)


def format_existing_workbook(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        format_worksheet(ws)
    wb.save(path)
