from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "outputs" / "julho_2025_mais_vendidos"
OUTPUT_JSON = OUTPUT_DIR / "julho_2025_mais_vendidos_data.json"


def money(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        text = str(value).strip()
        return float(text.replace(".", "").replace(",", ".") if "," in text else text)
    except Exception:
        return 0.0


def pick(row: dict[str, Any], *names: str, default=None):
    for name in names:
        value = row.get(name)
        if value not in (None, ""):
            return value
    return default


def reference_from_sku(sku: Any) -> str:
    text = str(sku or "").strip()
    if "-" in text:
        return text.split("-", 1)[1]
    return text


def find_july_workbook() -> Path:
    base = Path.home() / "OneDrive" / "Documentos" / "Clientologia"
    matches = [
        path
        for path in base.rglob("Base An*lise Magazord - Fechamento Julho.xlsx")
        if "\\2025\\" in str(path)
    ]
    if not matches:
        matches = [
            path
            for path in base.rglob("Base Análise Magazord - Fechamento Julho.xlsx")
            if "\\2025\\" in str(path)
        ]
    if not matches:
        raise FileNotFoundError("Não encontrei Base Análise Magazord - Fechamento Julho.xlsx em 2025.")
    return matches[0]


def rows_by_header(ws) -> list[dict[str, Any]]:
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {str(header): value for header, value in zip(headers, values) if header not in (None, "")}
        if any(value not in (None, "") for value in row.values()):
            rows.append(row)
    return rows


def build_price_lookup(price_rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    lookup: dict[str, dict[str, Any]] = {}
    for row in price_rows:
        sku = str(pick(row, "Produto Código", default="") or "").strip()
        if not sku:
            continue
        table_name = str(pick(row, "Tabela de Preços Nome", default="") or "").strip().lower()
        marketplace_id = money(pick(row, "Marketplace Id", default=0))
        rank = (0 if table_name == "padrão" else 1, 0 if marketplace_id == 0 else 1)
        old_price = money(pick(row, "Preço Antigo"))
        current = lookup.get(sku)
        if current and current["_rank"] <= rank:
            continue
        lookup[sku] = {
            "preco_antigo": old_price,
            "produto_nome": pick(row, "Produto Nome"),
            "categoria_preco": pick(row, "Produto Categoria"),
            "tabela_preco": pick(row, "Tabela de Preços Nome"),
            "_rank": rank,
        }
    return lookup


def main() -> None:
    source = find_july_workbook()
    wb = load_workbook(source, read_only=True, data_only=True)
    sales_rows = rows_by_header(wb["Vendas magazord - dinâmica"])
    price_rows = rows_by_header(wb["Tabela Preço Cheio"])
    price_lookup = build_price_lookup(price_rows)

    detail_rows: list[dict[str, Any]] = []
    grouped: dict[tuple[str, str, str], dict[str, Any]] = {}
    missing_old_price = 0

    for row in sales_rows:
        sku = str(pick(row, "Produto Código", default="") or "").strip()
        qty = money(pick(row, "Quantidade Item"))
        if not sku or qty <= 0:
            continue
        collection = pick(row, "Caracteristica", "Coleção", default="Sem coleção")
        category = pick(row, "Categoria Principal", default="")
        sold_value = money(pick(row, "Valor Total"))
        sale_unit_price = money(pick(row, "Valor Unitário"))
        price_info = price_lookup.get(sku, {})
        old_price = money(price_info.get("preco_antigo"))
        full_price_unit = old_price if old_price > 0 else sale_unit_price
        if old_price <= 0:
            missing_old_price += 1
        full_price_total = full_price_unit * qty
        discount = 1 - sold_value / full_price_total if full_price_total > 0 else 0
        reference = reference_from_sku(sku)
        product_name = price_info.get("produto_nome") or f"Ref. {reference}"
        key = (reference, str(product_name), str(collection or "Sem coleção"))

        detail_rows.append(
            {
                "SKU": sku,
                "Referência": reference,
                "Produto": product_name,
                "Coleção": collection,
                "Categoria": category,
                "Quantidade Vendida": qty,
                "Valor Vendido": sold_value,
                "Preço Venda Unit.": sale_unit_price,
                "Preço Antigo": old_price,
                "Preço Cheio Regra": full_price_unit,
                "Preço Cheio Total": full_price_total,
                "Desconto Vendido": discount,
                "Fonte Preço": "Preço Antigo histórico" if old_price > 0 else "Fallback Preço Venda",
            }
        )

        item = grouped.setdefault(
            key,
            {
                "Referência": reference,
                "Produto": product_name,
                "Coleção": collection,
                "Categoria Principal": category,
                "Quantidade Vendida": 0.0,
                "Valor Vendido": 0.0,
                "Preço Cheio Total": 0.0,
                "SKUs Vendidos": set(),
                "Preço Cheio Médio": 0.0,
                "Preço Médio Vendido": 0.0,
                "Desconto Vendido": 0.0,
            },
        )
        item["Quantidade Vendida"] += qty
        item["Valor Vendido"] += sold_value
        item["Preço Cheio Total"] += full_price_total
        item["SKUs Vendidos"].add(sku)

    product_rows: list[dict[str, Any]] = []
    for item in grouped.values():
        qty = money(item["Quantidade Vendida"])
        sold = money(item["Valor Vendido"])
        full = money(item["Preço Cheio Total"])
        product_rows.append(
            {
                "Referência": item["Referência"],
                "Produto": item["Produto"],
                "Coleção": item["Coleção"],
                "Categoria Principal": item["Categoria Principal"],
                "Quantidade Vendida": qty,
                "Valor Vendido": sold,
                "Preço Cheio Total": full,
                "Preço Médio Vendido": sold / qty if qty else 0,
                "Preço Cheio Médio": full / qty if qty else 0,
                "Desconto Vendido": 1 - sold / full if full else 0,
                "SKUs Vendidos": len(item["SKUs Vendidos"]),
            }
        )

    product_rows.sort(key=lambda item: (item["Quantidade Vendida"], item["Valor Vendido"]), reverse=True)
    detail_rows.sort(key=lambda item: (item["Quantidade Vendida"], item["Valor Vendido"]), reverse=True)

    collection_summary = []
    by_collection = defaultdict(lambda: {"Quantidade Vendida": 0.0, "Valor Vendido": 0.0, "Preço Cheio Total": 0.0, "Produtos": set()})
    for row in product_rows:
        bucket = by_collection[str(row["Coleção"])]
        bucket["Quantidade Vendida"] += money(row["Quantidade Vendida"])
        bucket["Valor Vendido"] += money(row["Valor Vendido"])
        bucket["Preço Cheio Total"] += money(row["Preço Cheio Total"])
        bucket["Produtos"].add(row["Referência"])
    for collection, bucket in by_collection.items():
        collection_summary.append(
            {
                "Coleção": collection,
                "Produtos": len(bucket["Produtos"]),
                "Quantidade Vendida": bucket["Quantidade Vendida"],
                "Valor Vendido": bucket["Valor Vendido"],
                "Preço Cheio Total": bucket["Preço Cheio Total"],
                "Desconto Vendido": 1 - bucket["Valor Vendido"] / bucket["Preço Cheio Total"] if bucket["Preço Cheio Total"] else 0,
            }
        )
    collection_summary.sort(key=lambda item: item["Quantidade Vendida"], reverse=True)

    summary = {
        "data_geracao": datetime.now().isoformat(timespec="seconds"),
        "fonte": str(source),
        "periodo": "Julho/2025",
        "linhas_venda": len(detail_rows),
        "produtos_referencias": len(product_rows),
        "quantidade_total": sum(money(row["Quantidade Vendida"]) for row in product_rows),
        "valor_vendido_total": sum(money(row["Valor Vendido"]) for row in product_rows),
        "preco_cheio_total": sum(money(row["Preço Cheio Total"]) for row in product_rows),
        "desconto_medio": 1 - sum(money(row["Valor Vendido"]) for row in product_rows) / sum(money(row["Preço Cheio Total"]) for row in product_rows),
        "linhas_com_fallback_preco_venda": missing_old_price,
        "regra_desconto": "Preço cheio = Preço Antigo quando maior que zero; senão Preço Venda. Desconto = 1 - Valor Vendido / Preço Cheio Total.",
    }

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(
        json.dumps(
            {
                "summary": summary,
                "products": product_rows,
                "details": detail_rows,
                "collections": collection_summary,
            },
            ensure_ascii=False,
            indent=2,
            default=list,
        ),
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
